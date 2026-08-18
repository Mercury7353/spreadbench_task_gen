import shutil
from pathlib import Path

from openpyxl import load_workbook

from spreadsheet_eval.verifier import norm, workbook_structure


def test_numeric_normalization_is_two_decimal_places():
    assert norm(1.234) == 1.23
    assert norm(1.235) == 1.24
    assert norm(True) is True


def test_empty_extra_sheet_is_a_structure_mismatch(tmp_path):
    source = next(
        (Path(__file__).resolve().parents[1] / "SpreadsheetEval").glob(
            "*/tests/input.xlsx"
        )
    )
    candidate = tmp_path / "candidate.xlsx"
    shutil.copy2(source, candidate)
    workbook = load_workbook(candidate)
    workbook.create_sheet("Unauthorized empty sheet")
    workbook.save(candidate)
    assert workbook_structure(source) != workbook_structure(candidate)
