import json
from pathlib import Path

from openpyxl import load_workbook

from spreadsheet_eval.provenance import canonical_workbook_digest, validate_provenance


def test_dataset_shape():
    root = Path(__file__).resolve().parents[1] / "SpreadsheetEval"
    tasks = sorted(p for p in root.iterdir() if p.is_dir())
    assert len(tasks) == 10
    assert sum(p.name.startswith("financial-") for p in tasks) == 5
    assert sum(p.name.startswith("debugging-") for p in tasks) == 5


def test_harbor_layout_and_oracle_payloads():
    root = Path(__file__).resolve().parents[1] / "SpreadsheetEval"
    required = [
        "instruction.md",
        "task.toml",
        "environment/Dockerfile",
        "environment/input.xlsx",
        "solution/solve.sh",
        "solution/reference.xlsx",
        "tests/test.sh",
        "tests/verifier.py",
        "tests/manifest.json",
        "tests/input.xlsx",
        "tests/reference.xlsx",
    ]
    for task in sorted(path for path in root.iterdir() if path.is_dir()):
        assert all((task / relative).is_file() for relative in required), task.name
        assert (task / "solution/solve.sh").stat().st_mode & 0o111
        assert (task / "tests/test.sh").stat().st_mode & 0o111
        assert canonical_workbook_digest(
            task / "environment/input.xlsx"
        ) == canonical_workbook_digest(task / "tests/input.xlsx")
        assert canonical_workbook_digest(
            task / "solution/reference.xlsx"
        ) == canonical_workbook_digest(task / "tests/reference.xlsx")


def test_required_cells_are_exactly_changed():
    root = Path(__file__).resolve().parents[1] / "SpreadsheetEval"
    for task in root.iterdir():
        if not task.is_dir():
            continue
        manifest = json.loads((task / "tests/manifest.json").read_text())
        before = load_workbook(task / "tests/input.xlsx", data_only=False)
        after = load_workbook(task / "tests/reference.xlsx", data_only=False)
        changed = set()
        for sheet in after.sheetnames:
            for row in after[sheet].iter_rows():
                for cell in row:
                    if before[sheet][cell.coordinate].value != cell.value:
                        changed.add(f"{sheet}!{cell.coordinate}")
        assert changed == set(manifest["required_cells"]), task.name


def test_every_target_weakness_is_grounded_in_original_cases():
    root = Path(__file__).resolve().parents[1]
    errors = validate_provenance(
        root / "SpreadsheetEval", root / "reports/original_template_97_cases.json"
    )
    assert not errors, "\n".join(errors)
