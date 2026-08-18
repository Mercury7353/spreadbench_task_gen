from openpyxl import Workbook

from spreadsheet_eval.mutation_operators import (
    blank_formula,
    choose_across_sheets,
    inject_formula_fault,
    mutate_formula,
)


def workbook_with_formulas():
    workbook = Workbook()
    workbook.active.title = "Model"
    workbook.create_sheet("Debt")
    for sheet in workbook.worksheets:
        for row in range(1, 21):
            sheet.cell(row, 1, f"=B{row}+1")
    return workbook


def test_selection_is_deterministic_and_cross_sheet():
    workbook = workbook_with_formulas()
    cells = {
        sheet.title: [cell for row in sheet.iter_rows() for cell in row if cell.value]
        for sheet in workbook.worksheets
    }
    first = choose_across_sheets(cells, 16, seed=7)
    second = choose_across_sheets(cells, 16, seed=7)
    assert first == second
    assert {sheet for sheet, _ in first} == {"Model", "Debt"}


def test_formula_mutations_are_auditable():
    workbook = workbook_with_formulas()
    blank = blank_formula(workbook, "Model", "A1")
    assert blank["reference"] == "=B1+1"
    assert workbook["Model"]["A1"].value is None

    fault = inject_formula_fault(workbook, "Debt", "A1", "reference_drift")
    assert fault["reference"] == "=B1+1"
    assert fault["injected"] == "=B2+1"
    assert mutate_formula("=SUM(A1:A3)", "sign_flip") == "=-1*(SUM(A1:A3))"
    assert mutate_formula("=A1", "offset") == "=1+(A1)"
