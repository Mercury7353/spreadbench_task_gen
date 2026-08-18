from __future__ import annotations

import json
import math
import shutil
import subprocess
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

TESTS = Path("/tests")
WORKSPACE = Path("/workspace")
REWARD = Path("/logs/verifier/reward.txt")


def recalc(source: Path, destination: Path) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        local = tmpdir / source.name
        outdir = tmpdir / "out"
        outdir.mkdir()
        shutil.copy2(source, local)
        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(outdir),
                str(local),
            ],
            check=True,
            capture_output=True,
        )
        converted = outdir / local.name
        shutil.copy2(converted, destination)


def raw_map(path: Path) -> dict[str, object]:
    wb = load_workbook(path, data_only=False)
    result = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    result[f"{ws.title}!{cell.coordinate}"] = cell.value
    return result


def workbook_structure(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=False)
    return [
        {
            "title": sheet.title,
            "state": sheet.sheet_state,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "freeze_panes": str(sheet.freeze_panes or ""),
            "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
        }
        for sheet in workbook.worksheets
    ]


def norm(value):
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return str(value)
        return round(float(value), 2)
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def values(path: Path) -> dict[str, object]:
    wb = load_workbook(path, data_only=True)
    return {
        f"{ws.title}!{c.coordinate}": norm(c.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
    }


def main() -> int:
    REWARD.parent.mkdir(parents=True, exist_ok=True)
    output = WORKSPACE / "output.xlsx"
    if not output.exists():
        print(json.dumps({"reason": "output_missing"}, indent=2))
        REWARD.write_text("0")
        return 0
    manifest = json.loads((TESTS / "manifest.json").read_text())
    expected_structure = workbook_structure(TESTS / "input.xlsx")
    actual_structure = workbook_structure(output)
    if actual_structure != expected_structure:
        print(
            json.dumps(
                {
                    "reason": "workbook_structure_mismatch",
                    "expected": expected_structure,
                    "actual": actual_structure,
                },
                indent=2,
            )
        )
        REWARD.write_text("0")
        return 0
    before, after = raw_map(TESTS / "input.xlsx"), raw_map(output)
    all_keys = set(before) | set(after)
    changed = {k for k in all_keys if before.get(k) != after.get(k)}
    required = set(manifest["required_cells"])
    if changed != required:
        print(
            json.dumps(
                {
                    "reason": "modified_set_mismatch",
                    "missing": sorted(required - changed),
                    "extra": sorted(changed - required),
                },
                indent=2,
            )
        )
        REWARD.write_text("0")
        return 0
    with tempfile.TemporaryDirectory() as tmp:
        actual, expected = Path(tmp) / "actual.xlsx", Path(tmp) / "expected.xlsx"
        recalc(output, actual)
        recalc(TESTS / "reference.xlsx", expected)
        av, ev = values(actual), values(expected)
        errors = {
            k: {"actual": av.get(k), "expected": ev.get(k)}
            for k in required
            if av.get(k) != ev.get(k)
        }
    reward = 0 if errors else 1
    print(
        json.dumps(
            {"reward": reward, "required_edits": len(required), "value_errors": errors},
            indent=2,
            default=str,
        )
    )
    REWARD.write_text(str(reward))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
