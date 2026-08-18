from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


def canonical_workbook_digest(path: Path) -> str:
    """Hash workbook semantics without ZIP timestamps or document metadata."""
    workbook = load_workbook(path, data_only=False, keep_links=True)
    digest = hashlib.sha256()
    for worksheet in workbook.worksheets:
        digest.update(json.dumps([worksheet.title, worksheet.sheet_state]).encode())
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None and not cell.has_style:
                    continue
                payload = [
                    cell.coordinate,
                    cell.data_type,
                    str(cell.value),
                    cell.number_format,
                    cell.style_id,
                ]
                digest.update(json.dumps(payload, ensure_ascii=False).encode())
    return digest.hexdigest()


def canonical_dataset_digests(dataset_root: Path) -> dict[str, str]:
    return {
        str(path.relative_to(dataset_root)): canonical_workbook_digest(path)
        for path in sorted(dataset_root.rglob("*.xlsx"))
    }


def validate_provenance(dataset_root: Path, original_report: Path) -> list[str]:
    """Return human-readable provenance errors for every generated task."""
    original = {
        row["case_id"]: sorted(row["weaknesses"])
        for row in json.loads(original_report.read_text())
    }
    errors: list[str] = []
    source_counts: dict[str, int] = {}
    for task in sorted(path for path in dataset_root.iterdir() if path.is_dir()):
        manifest = json.loads((task / "tests/manifest.json").read_text())
        task_id = manifest.get("task_id", task.name)
        source = manifest.get("source", {})
        filename = source.get("filename", "")
        source_counts[filename] = source_counts.get(filename, 0) + 1
        if source.get("license") not in {"Apache-2.0", "MIT"}:
            errors.append(f"{task_id}: source license is not approved")

        embedded = manifest.get("source_case_weaknesses", {})
        union: set[str] = set()
        for case_id in manifest.get("source_original_cases", []):
            if case_id not in original:
                errors.append(f"{task_id}: unknown original case {case_id}")
                continue
            union.update(original[case_id])
            if sorted(embedded.get(case_id, [])) != original[case_id]:
                errors.append(f"{task_id}: stale weakness evidence for {case_id}")
        ungrounded = set(manifest.get("weaknesses", [])) - union
        if ungrounded:
            errors.append(
                f"{task_id}: ungrounded target weaknesses {sorted(ungrounded)}"
            )

    if len(source_counts) != 5 or set(source_counts.values()) != {2}:
        errors.append(
            "dataset must contain exactly five distinct real sources used twice each"
        )
    return errors
