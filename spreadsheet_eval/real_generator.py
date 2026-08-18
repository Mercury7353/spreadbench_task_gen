from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .mutation_operators import (
    blank_formula,
    choose_across_sheets,
    inject_formula_fault,
)


@dataclass(frozen=True)
class RealSource:
    slug: str
    filename: str
    origin: str
    license: str


SOURCES = [
    RealSource(
        "manufacturing-plan",
        "5 Year Financial Plan for Manufacturing Company.xlsx",
        "IanMadlenya/finance-excel",
        "Apache-2.0",
    ),
    RealSource(
        "commercial-real-estate",
        "Commercial Real Estate Valuation Model.xlsx",
        "IanMadlenya/finance-excel",
        "Apache-2.0",
    ),
    RealSource(
        "coal-india",
        "Coal India Limited Financial Model & Valuation.xlsx",
        "IanMadlenya/finance-excel",
        "Apache-2.0",
    ),
    RealSource(
        "financial-projection",
        "Financial Projection Model.xlsx",
        "IanMadlenya/finance-excel",
        "Apache-2.0",
    ),
    RealSource(
        "project-finance",
        "Packt_Project_Finance.xlsx",
        "PacktPublishing/Project-Finance-and-Excel---Build-Financial-Models-from-Scratch",
        "MIT",
    ),
]

FINANCIAL_CASES = [
    ["04_04", "06_01"],
    ["06_02", "06_03"],
    ["06_04", "06_05"],
    ["07_01", "07_02"],
    ["11_01", "11_02"],
]
DEBUG_CASES = [
    ["01_02", "01_04"],
    ["02_02", "02_05"],
    ["03_01", "03_03"],
    ["05_01", "05_02"],
    ["06_06", "06_08"],
]
FINANCIAL_WEAKNESSES = [
    ["zero-target-coverage"],
    ["zero-target-coverage"],
    ["zero-target-coverage", "incomplete-target-coverage"],
    ["zero-target-coverage", "no-deliverable", "tool-protocol"],
    ["zero-target-coverage"],
]
DEBUG_WEAKNESSES = [
    [
        "incomplete-target-coverage",
        "over-editing",
        "numerical-looping",
        "context-bloat",
    ],
    ["incomplete-target-coverage", "over-editing"],
    ["incomplete-target-coverage", "over-editing"],
    ["incomplete-target-coverage", "over-editing", "exactness-near-miss"],
    ["incomplete-target-coverage", "over-editing", "context-bloat"],
]

ORIGINAL_CASE_WEAKNESSES = {
    "01_02": ["incomplete-target-coverage", "numerical-looping", "over-editing"],
    "01_04": [
        "context-bloat",
        "incomplete-target-coverage",
        "numerical-looping",
        "over-editing",
    ],
    "02_02": ["incomplete-target-coverage", "over-editing"],
    "02_05": ["incomplete-target-coverage", "over-editing"],
    "03_01": ["incomplete-target-coverage", "over-editing"],
    "03_03": ["incomplete-target-coverage", "over-editing"],
    "04_04": ["zero-target-coverage"],
    "05_01": ["exactness-near-miss", "incomplete-target-coverage", "over-editing"],
    "05_02": ["exactness-near-miss", "incomplete-target-coverage"],
    "06_01": ["zero-target-coverage"],
    "06_02": ["zero-target-coverage"],
    "06_03": ["zero-target-coverage"],
    "06_04": ["incomplete-target-coverage"],
    "06_05": ["zero-target-coverage"],
    "06_06": ["incomplete-target-coverage", "over-editing"],
    "06_08": ["context-bloat", "over-editing", "zero-target-coverage"],
    "07_01": ["zero-target-coverage"],
    "07_02": ["no-deliverable", "tool-protocol", "zero-target-coverage"],
    "11_01": ["zero-target-coverage"],
    "11_02": ["zero-target-coverage"],
}


def formula_cells(wb):
    by_sheet = {}
    for ws in wb.worksheets:
        cells = [
            c
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        if cells:
            by_sheet[ws.title] = cells
    return by_sheet


class RealWorkbookBuilder:
    def __init__(self, output: Path, source_dir: Path):
        self.output = output
        self.source_dir = source_dir

    def build(self):
        if self.output.exists():
            shutil.rmtree(self.output)
        self.output.mkdir(parents=True)
        summaries = []
        for i, source in enumerate(SOURCES, 1):
            src = self.source_dir / source.filename
            summaries.append(
                self._build_task(source, src, "financial", i, FINANCIAL_CASES[i - 1])
            )
            summaries.append(
                self._build_task(source, src, "debugging", i, DEBUG_CASES[i - 1])
            )
        (self.output / "dataset_manifest.json").write_text(
            json.dumps(summaries, indent=2) + "\n"
        )
        return summaries

    def _build_task(self, source, src, category, index, source_cases):
        task_id = f"{category}-{index:02d}-{source.slug}"
        task = self.output / task_id
        for sub in ["environment", "solution", "tests"]:
            (task / sub).mkdir(parents=True, exist_ok=True)
        reference = task / "tests/reference.xlsx"
        stage = task / "tests/.normalized-stage.xlsx"
        reference_wb = load_workbook(src, data_only=False, keep_links=True)
        reference_wb.save(stage)
        # A second round-trip makes conversion of legacy shared/array formulas
        # stable before the task-specific mutation is applied.
        stable_wb = load_workbook(stage, data_only=False, keep_links=True)
        stable_wb.save(reference)
        stage.unlink()
        # Derive input from the normalized reference. Independent round-trips
        # of genuine workbooks can normalize array/shared formulas differently,
        # which would otherwise create untracked changes.
        wb = load_workbook(reference, data_only=False, keep_links=True)
        by_sheet = formula_cells(wb)
        available = sum(map(len, by_sheet.values()))
        if category == "financial":
            selected = choose_across_sheets(
                by_sheet, min(140, max(60, available // 12)), 4100 + index
            )
            mutations = [blank_formula(wb, sheet, coord) for sheet, coord in selected]
            weakness = FINANCIAL_WEAKNESSES[index - 1]
            instruction = f"Complete every missing formula in the provided {source.slug} financial model. Preserve all existing values, formulas, formatting, sheet names, and workbook structure. Save the completed workbook as /workspace/output.xlsx."
        else:
            selected = choose_across_sheets(
                by_sheet, min(32, max(24, available // 70)), 7200 + index
            )
            kinds = ["sign_flip", "offset", "reference_drift"]
            mutations = []
            for n, (sheet, coord) in enumerate(selected):
                kind = kinds[n % len(kinds)]
                mutations.append(inject_formula_fault(wb, sheet, coord, kind))
            weakness = DEBUG_WEAKNESSES[index - 1]
            instruction = f"Audit and repair all incorrect formulas in the provided {source.slug} financial model. Change only defective cells; preserve every correct value, formula, formatting element, sheet name, and workbook structure. Save the repaired workbook as /workspace/output.xlsx."
        input_path = task / "tests/input.xlsx"
        wb.save(input_path)
        shutil.copy2(input_path, task / "environment/input.xlsx")
        shutil.copy2(reference, task / "solution/reference.xlsx")
        required = [f"{m['sheet']}!{m['cell']}" for m in mutations]
        manifest = {
            "task_id": task_id,
            "category": category,
            "source": source.__dict__,
            "source_original_cases": source_cases,
            "weaknesses": weakness,
            "source_case_weaknesses": {
                case_id: ORIGINAL_CASE_WEAKNESSES[case_id] for case_id in source_cases
            },
            "required_cells": required,
            "mutations": mutations,
        }
        (task / "tests/manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")
        (task / "instruction.md").write_text(instruction + "\n")
        (task / "task.toml").write_text(task_toml(task_id, category, len(required)))
        (task / "environment/Dockerfile").write_text(dockerfile())
        (task / "solution/solve.sh").write_text(
            "#!/bin/bash\nset -euo pipefail\ncp /solution/reference.xlsx /workspace/output.xlsx\n"
        )
        verifier_src = Path(__file__).with_name("verifier.py")
        shutil.copy2(verifier_src, task / "tests/verifier.py")
        (task / "tests/test.sh").write_text(
            "#!/bin/bash\nset -uo pipefail\nmkdir -p /logs/verifier\npython3 /tests/verifier.py\n"
        )
        for script in (task / "solution/solve.sh", task / "tests/test.sh"):
            script.chmod(0o755)
        return {
            "task_id": task_id,
            "category": category,
            "source": source.__dict__,
            "formula_count": available,
            "required_edits": len(required),
            "source_original_cases": source_cases,
            "weaknesses": weakness,
        }


def task_toml(task_id, category, edits):
    return f'''schema_version = "1.1"

[task]
name = "spreadsheet-eval/{task_id}"
description = "Real-workbook {category} task with {edits} exact edits"
authors = []
keywords = ["spreadsheet", "{category}", "real-workbook", "exact-match"]

[metadata]
category = "{category}"
difficulty = "hard"
tags = ["real-workbook", "strict-verifier", "weakness-grounded"]

[agent]
timeout_sec = 2400.0
[verifier]
timeout_sec = 300.0
[environment]
build_timeout_sec = 1800.0
cpus = 2
memory_mb = 4096
storage_mb = 8192
allow_internet = true
'''


def dockerfile():
    return """FROM python:3.11.10-bullseye
RUN sed -i -e 's|http://deb.debian.org|https://mirrors.aliyun.com|g' -e 's|http://security.debian.org|https://mirrors.aliyun.com/debian-security|g' /etc/apt/sources.list \\
 && apt-get -o Acquire::Retries=10 update \\
 && DEBIAN_FRONTEND=noninteractive apt-get -o Acquire::Retries=10 install -y --no-install-recommends libreoffice-calc python3-uno \\
 && rm -rf /var/lib/apt/lists/*
RUN pip install --no-cache-dir openpyxl==3.1.5
WORKDIR /workspace
COPY input.xlsx /workspace/input.xlsx
CMD ["sleep", "infinity"]
"""
